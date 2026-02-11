import pandas as pd
from fpdf import FPDF
import openpyxl
from datetime import datetime

class ExportManager:
    def __init__(self, calculator):
        self.calculator = calculator
        
    def format_currency(self, value):
        """Formata valor para o padrão brasileiro R$ 10.000,00"""
        try:
            return f"R$ {float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except (ValueError, TypeError):
            return "R$ 0,00"
        
    def export_to_excel(self, filename):
        """Exporta o orçamento para Excel"""
        try:
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            
            # Cria planilha resumida
            dados_resumo = self._prepare_resumo_data()
            df_resumo = pd.DataFrame(dados_resumo)
            df_resumo.to_excel(writer, sheet_name='ORÇAMENTO RESUMO', index=False)
            
            # Cria planilha detalhada
            dados_detalhados = self._prepare_detalhado_data()
            df_detalhado = pd.DataFrame(dados_detalhados)
            df_detalhado.to_excel(writer, sheet_name='DETALHAMENTO', index=False)
            
            # Cria planilha de frete
            dados_frete = self._prepare_frete_data()
            df_frete = pd.DataFrame(dados_frete)
            df_frete.to_excel(writer, sheet_name='FRETE', index=False)
            
            # Cria planilha de materiais
            dados_materiais = self._prepare_materiais_data()
            df_materiais = pd.DataFrame(dados_materiais)
            df_materiais.to_excel(writer, sheet_name='MATERIAIS', index=False)
            
            writer.close()
            return True
        except Exception as e:
            print(f"Erro ao exportar para Excel: {e}")
            return False
        
    def export_to_pdf(self, filename):
        """Exporta o orçamento para PDF"""
        try:
            pdf = FPDF()
            pdf.add_page()
            
            # Configurações
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, "ORÇAMENTO - GALPÃO PRÉ-MOLDADO", 0, 1, 'C')
            
            pdf.set_font("Arial", '', 12)
            pdf.ln(5)
            
            # Dados do cliente
            entrada = self.calculator.results['ENTRADA']
            pdf.cell(0, 8, f"Cliente: {entrada.get('C9', '')}", 0, 1)
            pdf.cell(0, 8, f"Telefone: {entrada.get('TELEFONE', '')}", 0, 1)
            pdf.cell(0, 8, f"Cidade: {entrada.get('C10', '')}", 0, 1)
            pdf.cell(0, 8, f"Data: {datetime.now().strftime('%d/%m/%Y')}", 0, 1)
            pdf.ln(5)
            
            # Dados do galpão
            pdf.cell(0, 8, f"Medida: {entrada.get('D2', '')}", 0, 1)
            pdf.cell(0, 8, f"Área: {entrada.get('D3', 0):.2f} m²", 0, 1)
            pdf.cell(0, 8, f"Telha: {entrada.get('C15', '')}", 0, 1)
            pdf.cell(0, 8, f"Cobertura: {entrada.get('C16', '')}", 0, 1)
            pdf.cell(0, 8, f"Com Nota: {entrada.get('COM_NOTA', 'NÃO')}", 0, 1)
            pdf.ln(5)
            
            # Itens do orçamento
            orcamento = self.calculator.get_orcamento_final()
            frete = self.calculator.results['FRETE']
            
            # Título da tabela
            pdf.set_font("Arial", 'B', 14)
            pdf.cell(0, 10, "RESUMO DO ORÇAMENTO", 0, 1, 'C')
            pdf.ln(5)
            
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(120, 10, "ITEM", 1)
            pdf.cell(60, 10, "VALOR", 1)
            pdf.ln()
            
            pdf.set_font("Arial", '', 11)
            itens = [
                ("COBERTURA (Telhas, Tesouras, Terças)", orcamento.get('custo_cobertura', 0)),
                ("ESTRUTURA (Pilares, Cálices, Vigas)", orcamento.get('custo_estrutura', 0)),
                ("COMPLEMENTOS", orcamento.get('custo_complementos', 0)),
                ("PROJETO", orcamento.get('custo_projeto', 0)),
                ("FRETE", orcamento.get('frete', 0)),
                ("CUSTO TOTAL + FRETE", orcamento.get('custo_total_com_frete', 0)),
                ("VALOR TOTAL DO ORÇAMENTO", orcamento.get('valor_venda', 0))
            ]
            
            for item, valor in itens:
                pdf.cell(120, 8, item, 1)
                
                if "VALOR TOTAL" in item:
                    pdf.set_font("Arial", 'B', 11)
                    pdf.cell(60, 8, self.format_currency(valor), 1)
                    pdf.set_font("Arial", '', 11)
                else:
                    pdf.cell(60, 8, self.format_currency(valor), 1)
                pdf.ln()
        
            # Informações do frete
            pdf.ln(8)
            pdf.set_font("Arial", 'B', 11)
            pdf.cell(0, 8, "INFORMAÇÕES DO FRETE:", 0, 1)
            pdf.set_font("Arial", '', 10)
            pdf.cell(0, 6, f"Distância: {frete.get('distancia', 0)} km | Total de viagens: {frete.get('total_viagens', 0)} | Valor por KM: R$ 2,00", 0, 1)
            
            # Custo do concreto
            custo_concreto_info = self.calculator.get_custo_concreto_info()
            pdf.ln(5)
            pdf.set_font("Arial", 'B', 11)
            pdf.cell(0, 8, "CUSTO DO CONCRETO:", 0, 1)
            pdf.set_font("Arial", '', 10)
            pdf.cell(0, 6, f"Valor por m³: {self.format_currency(custo_concreto_info['custo_por_m3'])}", 0, 1)
            
            # Observações
            pdf.ln(5)
            pdf.set_font("Arial", 'I', 10)
            pdf.cell(0, 6, "• Margem de lucro aplicada: 100% sobre o custo total + frete", 0, 1)
            
            if orcamento.get('com_nota', False):
                pdf.cell(0, 6, "• Inclui 8% para emissão de nota fiscal", 0, 1)
            
            # Detalhamento dos complementos
            pdf.ln(8)
            pdf.set_font("Arial", 'B', 11)
            pdf.cell(0, 8, "COMPLEMENTOS INCLUÍDOS:", 0, 1)
            pdf.set_font("Arial", '', 10)
            
            entrada = self.calculator.results['ENTRADA']
            complementos = [
                f"• Fechamento: {entrada.get('C11', 'NÃO')}",
                f"• Porta: {entrada.get('C12', 'NÃO')}",
                f"• Janela: {entrada.get('C13', 'NÃO')}",
                f"• Placas: {entrada.get('C14', 'NÃO')}",
                f"• Platibanda: {entrada.get('C17', 'NÃO')}",
                f"• Projeto: {entrada.get('C18', 'SIM')}",
                f"• Laje: {entrada.get('C19', 'NÃO')}",
                f"• Vigas: {entrada.get('C20', 'NÃO')}",
                f"• Portão: {entrada.get('C21', 'NÃO')}"
            ]
            
            for complemento in complementos:
                pdf.cell(0, 6, complemento, 0, 1)
            
            # Condições comerciais
            pdf.ln(10)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, "CONDIÇÕES COMERCIAIS:", 0, 1)
            pdf.set_font("Arial", '', 10)
            
            condicoes = [
                "PAGAMENTO:",
                "• 50% DE ENTRADA NA ASSINATURA DO CONTRATO;",
                "• 25% NO INICIO DA OBRA;", 
                "• 25% NA CONCLUSÃO DA OBRA;",
                "",
                "PRAZO:",
                "• 90 dias úteis após a assinatura do contrato;",
                "• Em caso de projeto, 90 dias após a entrega do projeto;",
                "",
                "NÃO INCLUI:",
                "• Sem piso, sem calha, sem fechamento e sem esquadrias;",
                "• Agregados e cimento para fundação por conta do contratante;",
                "• Se a fundação exceder 5 metros fica a custo do cliente;",
                "• Projeto de fundação (se necessário)."
            ]
            
            for condicao in condicoes:
                if condicao in ["PAGAMENTO:", "PRAZO:", "NÃO INCLUI:"]:
                    pdf.set_font("Arial", 'B', 10)
                    pdf.cell(0, 6, condicao, 0, 1)
                    pdf.set_font("Arial", '', 10)
                else:
                    pdf.cell(0, 6, condicao, 0, 1)
            
            # Rodapé
            pdf.ln(15)
            pdf.set_font("Arial", 'I', 8)
            pdf.cell(0, 5, f"Orçamento gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} - Válido por 10 dias", 0, 1, 'C')
            
            pdf.output(filename)
            return True
        except Exception as e:
            print(f"Erro ao exportar para PDF: {e}")
            return False
        
    def _prepare_resumo_data(self):
        """Prepara dados para o resumo"""
        orcamento = self.calculator.get_orcamento_final()
        entrada = self.calculator.results['ENTRADA']
        frete = self.calculator.results['FRETE']
        custo_concreto_info = self.calculator.get_custo_concreto_info()
        
        return {
            'Descrição': [
                'CLIENTE',
                'TELEFONE',
                'CIDADE',
                'MEDIDA DO GALPÃO', 
                'ÁREA TOTAL (m²)',
                'TIPO DE TELHA',
                'TIPO DE COBERTURA',
                'COM NOTA FISCAL',
                '---',
                'CUSTO COBERTURA',
                'CUSTO ESTRUTURA',
                'CUSTO COMPLEMENTOS',
                'CUSTO PROJETO',
                'CUSTO TOTAL MATERIAIS',
                '---',
                'CUSTO CONCRETO POR M³',
                '---',
                'DISTÂNCIA (km)',
                'TOTAL DE VIAGENS',
                'CUSTO FRETE',
                '---',
                'CUSTO TOTAL + FRETE',
                'MARGEM DE LUCRO (100%)',
                'VALOR TOTAL DO ORÇAMENTO'
            ],
            'Valor': [
                entrada.get('C9', ''),
                entrada.get('TELEFONE', ''),
                entrada.get('C10', ''),
                entrada.get('D2', ''),
                f"{entrada.get('D3', 0):.2f}",
                entrada.get('C15', ''),
                entrada.get('C16', ''),
                entrada.get('COM_NOTA', 'NÃO'),
                '',
                self.format_currency(orcamento.get('custo_cobertura', 0)),
                self.format_currency(orcamento.get('custo_estrutura', 0)),
                self.format_currency(orcamento.get('custo_complementos', 0)),
                self.format_currency(orcamento.get('custo_projeto', 0)),
                self.format_currency(orcamento.get('custo_total_materiais', 0)),
                '',
                self.format_currency(custo_concreto_info['custo_por_m3']),
                '',
                f"{frete.get('distancia', 0)}",
                f"{frete.get('total_viagens', 0)}",
                self.format_currency(orcamento.get('frete', 0)),
                '',
                self.format_currency(orcamento.get('custo_total_com_frete', 0)),
                '100%',
                self.format_currency(orcamento.get('valor_venda', 0))
            ]
        }
        
    def _prepare_detalhado_data(self):
        """Prepara dados detalhados"""
        entrada = self.calculator.results['ENTRADA']
        cobertura = self.calculator.results['COBERTURA']
        pilares = self.calculator.results['PILARES']
        calices = self.calculator.results['CALICES']
        vigas = self.calculator.results['VIGAS']
        fechamentos = self.calculator.results['FECHAMENTOS']
        laje = self.calculator.results['LAJE']
        portao = self.calculator.results['PORTAO']
        orcamento = self.calculator.get_orcamento_final()
        custo_concreto_info = self.calculator.get_custo_concreto_info()
        
        return {
            'Item': [
                'DADOS BÁSICOS',
                'Cliente',
                'Telefone', 
                'Cidade',
                'Medida',
                'Área (m²)',
                'Telha',
                'Cobertura',
                'Com Nota',
                '---',
                'COBERTURA',
                'Quantidade de Tesouras',
                'Custo Tesouras',
                'Metragem de Telha (m²)',
                'Custo Telhas',
                'Custo Terças',
                'Custo Eitão',
                'Custo Cumeeira',
                'Custo Contraventamento',
                '---',
                'ESTRUTURA',
                'Quantidade de Pilares',
                'Altura dos Pilares (m)',
                'Volume Concreto Pilares (m³)',
                'Custo Pilares',
                'Custo Cálices',
                'Metragem de Vigas (m)',
                'Custo Vigas',
                '---',
                'CUSTO DO CONCRETO',
                'Valor por m³',
                '---',
                'COMPLEMENTOS',
                'Fechamento',
                'Custo Fechamentos',
                'Platibanda',
                'Custo Platibanda',
                'Laje',
                'Custo Laje',
                'Portão',
                'Custo Portão',
                '---',
                'PROJETO',
                'Incluído',
                'Custo Projeto',
                '---',
                'FRETE',
                'Distância (km)',
                'Total de Viagens',
                'Valor do Frete',
                '---',
                'RESUMO FINAL',
                'Custo Total Materiais',
                'Custo Frete',
                'Custo Total + Frete',
                'Margem de Lucro',
                'Acréscimo Nota Fiscal',
                'VALOR TOTAL'
            ],
            'Valor/Detalhes': [
                '',
                entrada.get('C9', ''),
                entrada.get('TELEFONE', ''),
                entrada.get('C10', ''),
                entrada.get('D2', ''),
                f"{entrada.get('D3', 0):.2f}",
                entrada.get('C15', ''),
                entrada.get('C16', ''),
                entrada.get('COM_NOTA', 'NÃO'),
                '',
                '',
                f"{cobertura.get('C24', 0)}",
                self.format_currency(cobertura.get('C25', 0)),
                f"{cobertura.get('C30', 0):.1f}",
                self.format_currency(cobertura.get('C31', 0)),
                self.format_currency(cobertura.get('C40', 0)),
                self.format_currency(cobertura.get('C47', 0)),
                self.format_currency(cobertura.get('C52', 0)),
                self.format_currency(cobertura.get('CONTRAVENTAMENTO', 0)),
                '',
                '',
                f"{pilares.get('C11', 0)}",
                f"{pilares.get('C12', 0):.1f}",
                f"{pilares.get('volume_concreto', 0):.3f}",
                self.format_currency(pilares.get('C22', 0)),
                self.format_currency(calices.get('C11', 0)),
                f"{vigas.get('C12', 0):.1f}",
                self.format_currency(vigas.get('C22', 0)),
                '',
                '',
                self.format_currency(custo_concreto_info['custo_por_m3']),
                '',
                '',
                entrada.get('C11', 'NÃO'),
                self.format_currency(fechamentos.get('C9', 0)),
                entrada.get('C17', 'NÃO'),
                self.format_currency(fechamentos.get('PLATIBANDA', 0)),
                entrada.get('C19', 'NÃO'),
                self.format_currency(laje.get('B5', 0)),
                entrada.get('C21', 'NÃO'),
                self.format_currency(portao.get('C11', 0)),
                '',
                '',
                entrada.get('C18', 'SIM'),
                self.format_currency(orcamento.get('custo_projeto', 0)),
                '',
                '',
                f"{self.calculator.results['FRETE'].get('distancia', 0)}",
                f"{self.calculator.results['FRETE'].get('total_viagens', 0)}",
                self.format_currency(self.calculator.results['FRETE'].get('valor_final', 0)),
                '',
                '',
                self.format_currency(orcamento.get('custo_total_materiais', 0)),
                self.format_currency(orcamento.get('frete', 0)),
                self.format_currency(orcamento.get('custo_total_com_frete', 0)),
                '100%',
                '8%' if orcamento.get('com_nota', False) else '0%',
                self.format_currency(orcamento.get('valor_venda', 0))
            ]
        }
        
    def _prepare_frete_data(self):
        """Prepara dados do frete"""
        frete = self.calculator.results['FRETE']
        entrada = self.calculator.results['ENTRADA']
        
        return {
            'Item': [
                'Cidade de Destino',
                'Distância (km)',
                'Valor por KM',
                '---',
                'VIAGENS POR COMPONENTE',
                'Fundação',
                'Pilares',
                'Tesouras',
                'Placas',
                'Lajes',
                'Vigas',
                '---',
                'TOTAL DE VIAGENS',
                '---',
                'CÁLCULO DO FRETE',
                'Valor Ida',
                'Valor Ida e Volta',
                'Valor Final do Frete'
            ],
            'Valor': [
                entrada.get('C10', ''),
                f"{frete.get('distancia', 0)}",
                'R$ 2,00',
                '',
                '',
                '4',
                '3',
                '2',
                '3',
                '0',
                '0',
                '',
                f"{frete.get('total_viagens', 0)}",
                '',
                '',
                self.format_currency(frete.get('valor_ida', 0)),
                self.format_currency(frete.get('valor_ida_volta', 0)),
                self.format_currency(frete.get('valor_final', 0))
            ]
        }
        
    def _prepare_materiais_data(self):
        """Prepara dados dos materiais utilizados com medidas"""
        banco_dados = self.calculator.get_banco_dados()
        entrada = self.calculator.results['ENTRADA']
        
        # Filtra materiais relevantes baseados nas escolhas
        materiais_utilizados = []
        
        # Telha selecionada
        telha_selecionada = entrada.get('C15', 'TELHA SIMPLES')
        if telha_selecionada in banco_dados:
            item = banco_dados[telha_selecionada]
            materiais_utilizados.append(['Telha', telha_selecionada, 
                                       self.format_currency(item['valor']), item['medida']])
        
        # Perfis principais
        perfis_principais = [
            'PERFIL U 75mm 2,00',
            'PERFIL U 68mm 2,00', 
            'TERÇA ENR 75mm 2,00'
        ]
        
        for perfil in perfis_principais:
            if perfil in banco_dados:
                item = banco_dados[perfil]
                materiais_utilizados.append(['Perfil', perfil, 
                                           self.format_currency(item['valor']), item['medida']])
        
        # Ferros para estrutura
        ferros = [
            'FERRO_12.5',
            'FERRO_10', 
            'FERRO_5'
        ]
        
        for ferro in ferros:
            if ferro in banco_dados:
                item = banco_dados[ferro]
                nome_ferro = ferro.replace('_', ' ').title()
                materiais_utilizados.append(['Ferro', nome_ferro, 
                                           self.format_currency(item['valor']), item['medida']])
        
        # Concreto e agregados
        concretos = [
            'CIMENTO',
            'AREIA',
            'BRITA/PÓ',
            'CONCRETO_M3'
        ]
        
        for concreto in concretos:
            if concreto in banco_dados:
                item = banco_dados[concreto]
                materiais_utilizados.append(['Concreto/Agregado', concreto, 
                                           self.format_currency(item['valor']), item['medida']])
        
        # Acessórios
        acessorios = [
            'BARRA RED 1/4 VERGA',
            'CUMEEIRA',
            'CHAPA',
            'TELA_4.2'
        ]
        
        for acessorio in acessorios:
            if acessorio in banco_dados:
                item = banco_dados[acessorio]
                materiais_utilizados.append(['Acessório', acessorio, 
                                           self.format_currency(item['valor']), item['medida']])
        
        # Cria estrutura para DataFrame com 4 colunas
        dados_materiais = {
            'Categoria': [],
            'Material': [],
            'Valor Unitário': [],
            'Medida': []
        }
        
        for categoria, material, valor, medida in materiais_utilizados:
            dados_materiais['Categoria'].append(categoria)
            dados_materiais['Material'].append(material)
            dados_materiais['Valor Unitário'].append(valor)
            dados_materiais['Medida'].append(medida)
        
        return dados_materiais
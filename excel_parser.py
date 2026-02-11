import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries
import re

class ExcelParser:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.sheets_data = {}
        self.formulas = {}
        self.dependencies = {}
        
    def parse_excel(self):
        """Analisa o arquivo Excel e extrai dados e fórmulas"""
        wb = openpyxl.load_workbook(self.excel_file, data_only=False)
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['PDF COMPLEXO', 'PDF SIMPLES']:
                continue
                
            ws = wb[sheet_name]
            data = []
            sheet_formulas = {}
            
            # Extrai dados e fórmulas
            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    if cell.value is None:
                        row_data.append('')
                    elif isinstance(cell.value, str) and cell.value.startswith('='):
                        row_data.append(cell.value)
                        sheet_formulas[(cell.row, cell.column)] = cell.value
                    else:
                        row_data.append(cell.value)
                data.append(row_data)
            
            self.sheets_data[sheet_name] = pd.DataFrame(data)
            self.formulas[sheet_name] = sheet_formulas
            
        wb.close()
        return self.sheets_data, self.formulas
    
    def extract_cell_references(self, formula):
        """Extrai referências de células de uma fórmula"""
        # Padrões para referências de células (A1, B2, etc.)
        patterns = [
            r'[A-Z]+\d+',  # Referências simples
            r"'[^']*'![A-Z]+\d+",  # Referências entre planilhas
        ]
        
        references = []
        for pattern in patterns:
            matches = re.findall(pattern, formula)
            references.extend(matches)
            
        return references
    
    def build_dependency_graph(self):
        """Constrói grafo de dependências entre células"""
        dependencies = {}
        
        for sheet_name, formulas in self.formulas.items():
            for (row, col), formula in formulas.items():
                cell_ref = f"{sheet_name}!{openpyxl.utils.get_column_letter(col)}{row}"
                dependencies[cell_ref] = self.extract_cell_references(formula)
                
        return dependencies